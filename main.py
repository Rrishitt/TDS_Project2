from fastapi import FastAPI, Form, File, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import os
import openpyxl
from processing import fetch_answer
from typing import Optional

app = FastAPI()

# CORS Configuration (Vercel allows any origin by default)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "tasks.xlsx")

def load_tasks_from_excel():
    if not os.path.exists(EXCEL_FILE):
        return {}
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    tasks = {row[0]: row[1] for row in sheet.iter_rows(
        min_row=2, values_only=True) if row[0] and row[1]}
    
    tasks_answers = {row[0]: row[2] for row in sheet.iter_rows(
        min_row=2, values_only=True) if row[0] and row[2]}
    workbook.close()
    return (tasks, tasks_answers)

TASKS,TASKS_ANSWERS = load_tasks_from_excel()

def classify_task(question: str) -> str:
    """Classify a question based on keyword matching with TASKS."""
    question_lower = question.lower()  # Convert to lowercase for case-insensitive matching
    for task_id, keywords in TASKS.items():
        if any(keyword.lower() in question_lower for keyword in keywords.split(",")):
            return task_id  # Return the first matching task ID
    return "Unknown"  # Default if no match is found

def save_file(file: UploadFile):
    os.makedirs("uploads", exist_ok=True)
    if not file or not file.filename:
        return "Error: No file provided."
    # Define the file path
    file_path = os.path.join(os.getcwd(), "uploads", file.filename)
    try:
        # Write the file content manually
        with open(file_path, "wb") as buffer:
            buffer.write(file.file.read())
    except Exception as e:
        return f"Error saving file: {str(e)}"
    return file_path

@app.get("/", response_class=HTMLResponse)
async def serve_form():
    file_path = os.path.join(os.path.dirname(__file__), "index.html")
    try:
        with open(file_path, "r") as file:
            return HTMLResponse(content=file.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>index.html not found</h1>", status_code=404)

@app.post("/api/")
async def receive_question(
    question: str = Form(...),
    file: Optional[UploadFile] = File(None)
):
    try:
        task_id = classify_task(question)  # Ensure this function is implemented properly

        file_path = None
        if file:
            file_path = save_file(file)
            print(file_path)

        if task_id in {'GA1.3', 'GA1.4', 'GA1.5', 'GA1.7', 'GA1.8', 'GA1.9', 'GA1.10', 'GA1.12'}:
            answer = fetch_answer(task_id=task_id, question=question, file_path=file_path or "")
        else:
            answer = TASKS_ANSWERS.get(task_id, "No answer found for this task.")

        return {
            "question": question,
            "task": task_id,
            "answer": answer,
            "file_received": file.filename if file else "No file uploaded",
        }
    
    except HTTPException as e:
        # Handles specific HTTP exceptions
        return JSONResponse(status_code=e.status_code, content={"error": str(e)})
    
    except Exception as e:
        # General error handling for internal server errors (500)
        return JSONResponse(status_code=500, content={"error": "Internal Server Error", "details": str(e)})
